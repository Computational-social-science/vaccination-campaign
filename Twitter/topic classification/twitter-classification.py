import datetime
import numpy
import numpy as np
import xlrd, xlwt
import openpyxl
import pandas as pd
import time
import os
import re
import operator
import math
import operator

#f=open('F-C.csv','w')
def cipin(s,diction):
    s_dict = {}
    for letter in s:
        s_list=letter.split()
        for item in s_list:
            if item not in diction:
                if item not in s_dict:
                    s_dict[item]=1
                else:
                    s_dict[item]+=1

    return s_dict
if __name__ == '__main__':
    '''
    处理停用词
    '''
    stop_list=[]
    with open("C:/Users/mikij/Desktop/代码/推特程序/推特维度提取/stop.txt", "r",encoding='utf-8-sig') as f:
        for stop in f.readlines():
            stop_list.append(stop.strip('\n').strip())# 去掉列表中每一个元素的换行符
    #print(stop_list)

    '''
    创建结果写入文档
    '''
    f = xlwt.Workbook()  # 创建工作簿
    # sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    # i = 1
    # head = ['word','fre']  # 标题
    # for j in range(0, len(head)):  # 写标题
    #     sheet1.write(0, j, head[j])
    '''
    打开分类关键词文档
    '''
    keywords_dict={}
    wp_path='C:/Users/mikij/Desktop/代码/推特程序/推特维度提取/Twitter情感.xlsx'
    wp = xlrd.open_workbook(wp_path)
    sheets_wp = wp.sheet_names()
    for name in sheets_wp:

        data = pd.read_excel(wp_path, sheet_name=name)  # ,sheet_name=sheet_names[i]
        list1 = data.values.tolist()
        list1 = np.array(list1)
        for keywords in list1:
            keywords_dict.setdefault(name, []).append(keywords[0])
        # if 'trial' in keywords_dict[name]:
        #     print(keywords_dict[name])

    twitter_dict={}
    time_dict={}
    location_dict={}
    all = []
    path1='C:/Users/mikij/Desktop/代码/推特程序/推特维度提取/推特数据汇总'
    dirs1 = os.listdir(path1)
    for file in dirs1:
        print(file)
        intput_file = file
        path=path1 + '/' + intput_file
        wb = xlrd.open_workbook(path)
        sheets = wb.sheet_names()
       # data_xls = pd.ExcelFile(path)
       # sheet_name_list=data_xls.sheet_names
        for name in sheets:

            data = pd.read_excel(path,sheet_name=name)#,sheet_name=sheet_names[i]

            list1 = data.values.tolist()
            list1 = np.array(list1)
            for sentence in list1:
                for namee in sheets_wp:
                    for word in keywords_dict[namee]:
                        if word in sentence[0]:
                            twitter_dict.setdefault(namee, []).append(sentence[0])
                            time_dict.setdefault(namee, []).append(sentence[1])
                            location_dict.setdefault(namee, []).append(sentence[2])
                            #f.save('推特分类结果.xlsx')
                            break
    
    #save

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    for name in sheets_wp:
        outws = outwb.create_sheet(name)
        outws.cell(1,1,"twitter")
        outws.cell(1,2,"time")
        outws.cell(1,3,"location")
        dataframe = pd.DataFrame({'twitter':twitter_dict[name],'time':time_dict[name],'location':location_dict[name]})
       
        for i in range(dataframe.shape[0]):
            for j in range(dataframe.shape[1]):
                outws.cell(column = j+1 , row = i+2 , value = "%s" % dataframe.iloc[i,j])
    
    savexlsx = "./推特分类结果.xlsx"
    outwb.save(savexlsx)  # 保存结果

    
   ''' 
    with pd.ExcelWriter('推特分类结果.xls') as writer:
        for name in sheets_wp:
            dataframe = pd.DataFrame({'twitter':twitter_dict[name],'time':time_dict[name],'location':location_dict[name]})
            dataframe.to_excel(writer,sheet_name=name)
            
            for line in dataframe:
                for x in range(len(line)):
                    print(x)
        #print(dataframe)
    # sheet_chu = "sheet%s" % name
    # o = f.add_sheet(str(sheet_chu), cell_overwrite_ok=True)

   # print(len(all))
'''
