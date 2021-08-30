# -*- coding: utf-8 -*-
"""
Created on Tue Mar 23 17:25:46 2021

@author: mikij
"""
import os
os.chdir('C:/Users/mikij/Desktop/实验/代码/推特程序/推特维度提取/')

import numpy as np
import xlrd, xlwt
import openpyxl
import pandas as pd
import re
import emoji
import spacy

def sentence_clean(sentence):
    sentence = re.sub(r'[^\x00-\x7F]+','', sentence)
    sentence = re.sub(' rt ','', sentence)
    sentence = re.sub('(\.)+','.', sentence)
    sentence = re.sub('((www\.[^\s]+))','',sentence)
    sentence = re.sub('((http://[^\s]+))','',sentence)
    sentence = re.sub('((https://[^\s]+))','',sentence)
    sentence = re.sub('@[^\s]+','',sentence)
    sentence = re.sub('[\s]+', ' ', sentence)
    sentence = re.sub(r'#([^\s]+)', r'\1', sentence)
    sentence = re.sub('\$','',sentence)
    #sentence = re.sub('%','',sentence)
    sentence = re.sub('^','',sentence)
    sentence = re.sub('&','',sentence)
    sentence = re.sub('\*','',sentence)
    sentence = re.sub('\(','',sentence)
    sentence = re.sub('\)','',sentence)
    sentence = re.sub('-','',sentence)
    sentence = re.sub('\+','',sentence)
    sentence = re.sub('=','',sentence)
    sentence = re.sub('"','',sentence)
    sentence = re.sub('\t','',sentence)
    sentence = re.sub('\r','',sentence)
    sentence = re.sub('\n','',sentence)
    sentence = re.sub('~','',sentence)
    sentence = re.sub('`','',sentence)
    sentence = re.sub('^-?[0-9]+$','', sentence)
    #sentence = sentence.strip('\'"')
    return sentence

#处理停用词
stop_list=[]
with open("C:/Users/mikij/Desktop/实验/代码/推特程序/推特维度提取/stop.txt", "r",encoding='utf-8-sig') as f:
    for stop in f.readlines():
        stop_list.append(stop.strip('\n').strip())# 去掉列表中每一个元素的换行符
            
#打开分类关键词文档
keywords_dict={}
wp_path='C:/Users/mikij/Desktop/实验/代码/推特程序/推特维度提取/Twitter情感.xlsx'
wp = xlrd.open_workbook(wp_path)
sheets_wp = wp.sheet_names()
for name in sheets_wp:
    data = pd.read_excel(wp_path, sheet_name=name)  # ,sheet_name=sheet_names[i]
    list1 = data.values.tolist()
    list1 = np.array(list1)
    for keywords in list1:
        keywords_dict.setdefault(name, []).append(keywords[0])


file_path ="C:/Users/mikij/Desktop/实验/一些结果/各主题分类结果-0306"
# 存放所有文件名
file_list = []
# 存放每个子文件夹下所对应的文件名
file_dict = {}
for iroot, idirs, ifiles in os.walk(file_path):
    if not idirs:
        #ifiles.remove('.DS_Store')
        file_list.extend(ifiles)
        file_dict[iroot] = ifiles

twitter_dict={}
time_dict={}
location_dict={}
for file in file_list:
    print(file)
    path=os.path.join(file_path, file)
    file_data=pd.read_excel(path,encoding='utf-8')
    tw_list = file_data.values.tolist()
    tw_list = np.array(tw_list) #tw_list[0][0]=content
    for tw in tw_list: #每条推特
        sentence=tw[0]
        sentence = emoji.demojize(sentence)
        sentence_t = sentence_clean(sentence)       
        nlp = spacy.load("en_core_web_sm")
        sentences = nlp(sentence_t).sents
        for sent in sentences: #sent就是切分后的句子
            for namee in sheets_wp:
                    for word in keywords_dict[namee]:
                        sent=str(sent)
                        if word in sent:
                            twitter_dict.setdefault(namee, []).append(sent)
                            time_dict.setdefault(namee, []).append(tw[1])
                            location_dict.setdefault(namee, []).append(tw[2])
                            #f.save('推特分类结果.xlsx')
                            break
#save
save_path='C:/Users/mikij/Desktop/实验/代码/推特程序/推特维度提取/分句结果/'
for name in sheets_wp:
    outwb = openpyxl.Workbook()
    outws = outwb.active
    outws.cell(row=1,column =1,value="twitter")
    outws.cell(row=1,column =2,value="time")
    outws.cell(row=1,column =3,value="location")
    dataframe = pd.DataFrame({'twitter':twitter_dict[name],'time':time_dict[name],'location':location_dict[name]})
    
    for i in range(dataframe.shape[0]):
        for j in range(dataframe.shape[1]):
            outws.cell(column = j+1 , row = i+2 , value = "%s" % dataframe.iloc[i,j])
    path=os.path.join(save_path, name)    
    outwb.save(path + ".xlsx")