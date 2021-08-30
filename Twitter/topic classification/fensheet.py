# -*- coding: utf-8 -*-
"""
Created on Sat Jan 30 22:10:14 2021

@author: Administrator
"""

import os
os.chdir('C:/Users/Administrator/Desktop/推特维度提取/')


from openpyxl import load_workbook, Workbook
 
wb = load_workbook("test.xlsx")
sheetnames = wb.sheetnames
 
for name in sheetnames:
    ws = wb.get_sheet_by_name(name)
    print(ws)
    # 创建新的Excel
    wb2 = Workbook()
    # 获取当前sheet
    ws2 = wb2.active
    ws2.cell(1,1,'sentence')
    ws2.cell(1,2,'time')
    ws2.cell(1,3,'location')
    # 两个for循环遍历整个excel的单元格内容
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            # 写入新Excel
            
            ws2.cell(row=i + 2, column=j + 1, value=cell.value)
            # 设置新Sheet的名称
            ws2.title = name
 
    wb2.save(name + ".xlsx")
    